"""
PARADAIS DDB · Dashboard Ejecutivo — Streamlit Cloud
Versión autónoma, 100% gratuita, sin servidor local.
Seguridad: auth por hash SHA-256, datos solo en RAM de sesión,
           sin trazas de error expuestas, validación de archivo estricta.
"""

import io
import re
import time
import hashlib
import logging
from datetime import datetime
from typing import Optional

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import plotly.express as px
import plotly.graph_objects as go

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA  (debe ser el primer comando Streamlit)
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Paradais DDB · Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS RESPONSIVO (desktop + móvil)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Paleta Elegancia Editorial ──────────────────────────────────────────── */
:root {
    --bg:        #FDFCFA;
    --card:      #FFFFFF;
    --text:      #1C1917;
    --text-soft: #78716C;
    --accent:    #EAB308;
    --border:    #E7E5E4;
    --shadow:    rgba(0,0,0,0.03) 0px 4px 12px;
    --green:     #10B981;
    --amber:     #F59E0B;
    --red:       #DC2626;
    --gray-line: #475569;
}

/* Fondo canvas */
.stApp, [data-testid="stAppViewContainer"] {
    background-color: #FDFCFA !important;
}
[data-testid="stHeader"] { background: transparent !important; }

/* ── Header principal ─────────────────────────────────────────────────────── */
.ddb-header {
    display: flex;
    align-items: center;
    gap: 1rem;
    padding: 1rem 0 0.5rem 0;
    border-bottom: 3px solid #EAB308;
    margin-bottom: 1.4rem;
}
.ddb-logo {
    width: 48px; height: 48px;
    background: #000000;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
    overflow: hidden;
}
.ddb-logo svg { width: 48px; height: 48px; }
.ddb-title {
    font-family: Georgia, 'Times New Roman', serif;
    font-size: 1.55rem;
    font-weight: 700;
    color: #1C1917;
    line-height: 1.1;
}
.ddb-sub {
    font-size: 0.78rem;
    color: #78716C;
    margin-top: 0.1rem;
    font-family: Georgia, serif;
    font-style: italic;
}
.ddb-badge {
    margin-left: auto;
    background: #F0FDF4;
    border: 1px solid #BBF7D0;
    color: #15803D;
    font-size: 0.72rem;
    font-weight: 600;
    padding: 0.3rem 0.7rem;
    border-radius: 20px;
    white-space: nowrap;
}

/* ── Tarjetas KPI ─────────────────────────────────────────────────────────── */
.kpi-card {
    background: #FFFFFF;
    border: 1px solid #E7E5E4;
    border-radius: 14px;
    padding: 1.1rem 1.4rem 1rem 1.4rem;
    box-shadow: rgba(0,0,0,0.03) 0px 4px 12px;
    margin-bottom: 0.6rem;
    min-height: 108px;
    position: relative;
    overflow: hidden;
}
.kpi-card::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: #EAB308;
    border-radius: 14px 14px 0 0;
}
.kpi-label {
    font-size: 0.68rem;
    font-weight: 700;
    color: #78716C;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 0.4rem;
}
.kpi-value {
    font-size: 1.75rem;
    font-weight: 700;
    color: #1C1917;
    line-height: 1.1;
    font-family: Georgia, serif;
}
.kpi-badge {
    display: inline-block;
    font-size: 0.65rem;
    font-weight: 700;
    padding: 0.15rem 0.5rem;
    border-radius: 20px;
    margin-top: 0.3rem;
}
.badge-green  { background:#D1FAE5; color:#065F46; }
.badge-amber  { background:#FEF3C7; color:#92400E; }
.badge-red    { background:#FEE2E2; color:#991B1B; }
.badge-neutral{ background:#F5F5F4; color:#57534E; }

/* ── Semáforos mate ───────────────────────────────────────────────────────── */
.sem-verde {
    background: #F0FDF4;
    border-left: 4px solid #10B981;
    color: #1C1917;
    padding: 0.55rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 0.3rem 0;
    font-size: 0.86rem;
}
.sem-amber {
    background: #FFFBEB;
    border-left: 4px solid #F59E0B;
    color: #1C1917;
    padding: 0.55rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 0.3rem 0;
    font-size: 0.86rem;
}
.sem-rojo {
    background: #FEF2F2;
    border-left: 4px solid #DC2626;
    color: #1C1917;
    padding: 0.55rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 0.3rem 0;
    font-size: 0.86rem;
}

/* ── Títulos de sección ──────────────────────────────────────────────────── */
.section-title {
    font-family: Georgia, serif;
    font-size: 1rem;
    font-weight: 700;
    color: #1C1917;
    padding-bottom: 0.35rem;
    border-bottom: 2px solid #F5F5F4;
    margin: 1.4rem 0 0.75rem 0;
}

/* ── Bloque Banco del Pacífico ─────────────────────────────────────────────── */
.bp-block {
    background: #FFFFFF;
    border: 1px solid #E7E5E4;
    border-radius: 14px;
    padding: 1.2rem 1.4rem;
    box-shadow: rgba(0,0,0,0.03) 0px 4px 12px;
}
.bp-block-title {
    font-family: Georgia, serif;
    font-size: 1.05rem;
    font-weight: 700;
    color: #1C1917;
    margin-bottom: 0.8rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

/* ── Sidebar ─────────────────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background-color: #FDFCFA !important;
    border-right: 1px solid #E7E5E4;
}
[data-testid="stSidebar"] * { color: #1C1917 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stMultiSelect label { font-size: 0.8rem !important; }

/* ── Selectbox / Multiselect — texto siempre visible ────────────────────── */
/* Contenedor seleccionado */
[data-baseweb="select"] > div:first-child,
[data-baseweb="select"] [data-testid="stSelectboxVirtualDropdown"],
[data-baseweb="select"] input,
[data-baseweb="select"] span,
[data-baseweb="select"] div[class*="ValueContainer"] > div,
[data-baseweb="select"] div[class*="singleValue"],
[data-baseweb="select"] div[class*="placeholder"] {
    color: #1C1917 !important;
    background-color: #FFFFFF !important;
}
/* Caja exterior del select */
[data-baseweb="select"] > div {
    background-color: #FFFFFF !important;
    border-color: #E7E5E4 !important;
}
/* Opciones del dropdown */
[data-baseweb="popover"] li,
[data-baseweb="menu"] li,
[data-baseweb="menu"] [role="option"],
ul[data-testid="stSelectboxVirtualDropdown"] li {
    color: #1C1917 !important;
    background-color: #FFFFFF !important;
}
[data-baseweb="menu"] [role="option"]:hover,
[data-baseweb="menu"] [aria-selected="true"] {
    background-color: #FEF9EC !important;
    color: #1C1917 !important;
}
/* Multiselect tags */
[data-baseweb="tag"] {
    background-color: #FEF3C7 !important;
    color: #92400E !important;
}
[data-baseweb="tag"] span { color: #92400E !important; }
/* Input dentro del multiselect */
[data-baseweb="select"] input { color: #1C1917 !important; }
/* Label general de widgets */
.stSelectbox label, .stMultiSelect label,
[data-testid="stWidgetLabel"] { color: #1C1917 !important; }

/* ── Botones ─────────────────────────────────────────────────────────────── */
.stButton > button {
    border-radius: 8px;
    font-weight: 600;
    border: 1px solid #E7E5E4 !important;
    color: #1C1917 !important;
    background: #FFFFFF !important;
}
.stDownloadButton > button {
    background: #1C1917 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px;
    font-weight: 600;
}

/* ── Responsive móvil ────────────────────────────────────────────────────── */
@media (max-width: 768px) {
    [data-testid="column"] {
        width: 100% !important;
        flex: none !important;
        min-width: 100% !important;
    }
    .kpi-value  { font-size: 1.4rem; }
    .ddb-title  { font-size: 1.2rem; }
    .ddb-badge  { display: none; }
}

/* ── Date input — mismo estilo que selectbox ────────────────────────────── */
[data-testid="stDateInput"] input,
[data-testid="stDateInput"] > div > div,
[data-baseweb="input"] input,
[data-baseweb="calendar"] {
    background-color: #FFFFFF !important;
    color: #1C1917 !important;
    border-color: #E7E5E4 !important;
}
[data-baseweb="input"] {
    background-color: #FFFFFF !important;
    border-color: #E7E5E4 !important;
}
[data-baseweb="input"] > div {
    background-color: #FFFFFF !important;
}
[data-baseweb="input"] input {
    color: #1C1917 !important;
    background-color: #FFFFFF !important;
}
/* Calendario popup */
[data-baseweb="calendar"] * {
    background-color: #FFFFFF !important;
    color: #1C1917 !important;
}
[data-baseweb="calendar"] [aria-selected="true"] {
    background-color: #EAB308 !important;
    color: #FFFFFF !important;
}

/* ── Ocultar elementos de Streamlit ──────────────────────────────────────── */
#MainMenu, footer, [data-testid="stToolbar"] { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# AUTENTICACIÓN — contraseña hash SHA-256 desde st.secrets
# ─────────────────────────────────────────────────────────────────────────────
_MAX_INTENTOS  = 5
_BLOQUEO_SEGS  = 300   # 5 minutos


def _sha256(texto: str) -> str:
    return hashlib.sha256(texto.encode("utf-8")).hexdigest()


def _hash_esperado() -> str:
    """
    Lee el hash SHA-256 desde st.secrets['auth']['password_hash'].
    NUNCA se define una contraseña por defecto en el código.
    Si el secret no existe, la app se detiene con un mensaje de configuración.
    """
    try:
        h = st.secrets["auth"]["password_hash"]
        if not h or len(h) != 64:
            raise ValueError("Hash inválido")
        return h
    except Exception:
        st.error(
            "⚙️ **Configuración requerida:** "
            "El administrador debe configurar `[auth] password_hash` en los Secrets de esta app."
        )
        st.stop()


def pantalla_login() -> bool:
    """Muestra la pantalla de login. Retorna True solo cuando el usuario está autenticado."""
    if st.session_state.get("autenticado"):
        return True

    ahora = time.time()
    intentos      = st.session_state.get("_intentos", 0)
    bloqueado_hasta = st.session_state.get("_bloqueado_hasta", 0)

    if ahora < bloqueado_hasta:
        restante = int(bloqueado_hasta - ahora)
        st.error(f"🔒 Demasiados intentos fallidos. Intenta de nuevo en {restante} segundos.")
        st.stop()

    _, col_c, _ = st.columns([1, 1.6, 1])
    with col_c:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            "<h2 style='text-align:center;color:#e2e8f0'>📊 Paradais DDB</h2>"
            "<p style='text-align:center;color:#94a3b8'>Dashboard Ejecutivo · Ingresa tu contraseña</p>",
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        with st.form("login", clear_on_submit=True):
            pwd = st.text_input("Contraseña", type="password", placeholder="••••••••",
                                label_visibility="collapsed")
            ok = st.form_submit_button("Ingresar →", use_container_width=True, type="primary")

        if ok:
            if _sha256(pwd) == _hash_esperado():
                st.session_state["autenticado"] = True
                st.session_state["_intentos"]   = 0
                st.rerun()
            else:
                intentos += 1
                st.session_state["_intentos"] = intentos
                if intentos >= _MAX_INTENTOS:
                    st.session_state["_bloqueado_hasta"] = ahora + _BLOQUEO_SEGS
                    st.error(f"🔒 Cuenta bloqueada por {_BLOQUEO_SEGS // 60} minutos.")
                else:
                    quedan = _MAX_INTENTOS - intentos
                    st.error(f"❌ Contraseña incorrecta. Intentos restantes: {quedan}")

    return False


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES DEL PARSER
# ─────────────────────────────────────────────────────────────────────────────
MESES = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE",
]
MES_NUM = {m: i + 1 for i, m in enumerate(MESES)}

EXTRA_SHEETS = ["Paradais Media 2026"]

HEADER_KW = {
    "fecha","cliente","departamento","tipo","ciudad",
    "base iva","comision","costos","rentabilidad","periodo","costo real",
}

COL_ALIASES = {
    "fecha":          ["FECHA","Fecha","fecha","DATE","date"],
    "cliente":        ["CLIENTE","Cliente","cliente","ANUNCIANTE","Anunciante"],
    "departamento":   ["DEPARTAMENTO","Departamento","departamento","DEPTO","Depto"],
    "base_iva":       ["BASE IVA","Base IVA","Base Iva","BASE_IVA","BASEIVA",
                       "Ventas","VENTAS","NETO","Neto","BASE IVA 12%"],
    "comision":       ["COMISION","Comisión","COMISIÓN","Comision","COMIS","comision"],
    "costos":         ["COSTOS","Costos","costos","COSTO","Costo","COST",
                       "Costo Real","COSTO REAL","CostoReal"],
    "margen":         ["MARGEN","Margen","margen","UTILIDAD","Utilidad",
                       "Rentabilidad","RENTABILIDAD","MARGEN BRUTO"],
    "tipo":           ["TIPO","Tipo","tipo","TYPE","MODALIDAD","Modalidad"],
    "ciudad":         ["CIUDAD","Ciudad","ciudad","CITY"],
    "periodo_ingreso":["PERIODO INGRESO","Periodo Ingreso","PERIODO_INGRESO",
                       "PERIODO","Periodo","Periodo Ingreso"],
}

# Columnas de la hoja limpia exportada
EXPORT_COLS = [
    "fecha","empresa","mes","ano","cliente","departamento_limpio",
    "tipo","ciudad","base_iva","comision","total_venta_real","costos","margen",
]
EXPORT_HEADERS = [
    "Fecha","Empresa","Mes","Año","Cliente","Departamento",
    "Tipo","Ciudad","Base IVA","Comisión","Total Venta Real","Costos","Margen",
]


# ─────────────────────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────────────────────
def _normalizar_depto(valor) -> Optional[str]:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    d = str(valor).upper().strip()
    d = re.sub(r'\s+COMISI[OÓ]N(ES)?\s*$', '', d, flags=re.IGNORECASE)
    d = re.sub(r'^COMISI[OÓ]N(ES)?\s+', '', d, flags=re.IGNORECASE)
    d = " ".join(d.split())
    return d.strip() or None


def _mapear_columnas(cols: list) -> dict:
    upper = {c.strip().upper(): c for c in cols}
    mapa = {}
    for can, aliases in COL_ALIASES.items():
        for alias in aliases:
            if alias.upper() in upper:
                mapa[can] = upper[alias.upper()]
                break
    return mapa


def _fila_header(xl: pd.ExcelFile, hoja: str, max_scan: int = 15) -> int:
    try:
        raw = xl.parse(hoja, header=None, nrows=max_scan)
    except Exception:
        return 0
    for idx, fila in raw.iterrows():
        vals = {str(v).lower().strip() for v in fila.values if pd.notna(v) and str(v).strip()}
        if sum(1 for kw in HEADER_KW if any(kw in v for v in vals)) >= 2:
            return int(idx)
    return 0


def _leer_hoja(xl: pd.ExcelFile, hoja: str, empresa: str,
               mes: Optional[str] = None, ano: Optional[int] = None) -> pd.DataFrame:
    try:
        hr = _fila_header(xl, hoja)
        df = xl.parse(hoja, header=hr)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    mapa   = _mapear_columnas(list(df.columns))
    rename = {v: k for k, v in mapa.items()}
    df     = df.rename(columns=rename)

    for col_req in ("fecha", "cliente", "departamento"):
        if col_req not in df.columns:
            return pd.DataFrame()

    # Filtrar filas inválidas
    df = df[df["fecha"].notna()]
    df = df[df["departamento"].notna()]
    df = df[df["departamento"].astype(str).str.strip() != ""]

    if "cliente" in df.columns:
        df = df[df["cliente"].notna()]
        df = df[~df["cliente"].astype(str).str.upper().str.contains(
            r'\bTOTAL\b|\bSUBTOTAL\b|\bSUMA\b|\bTOTALES\b|\bCLIENTE\b',
            regex=True, na=False,
        )]

    # Eliminar filas-encabezado duplicadas
    for col, val in [("tipo","TIPO"), ("ciudad","CIUDAD"), ("departamento","DEPARTAMENTO")]:
        if col in df.columns:
            df = df[df[col].astype(str).str.upper().str.strip() != val]

    if df.empty:
        return pd.DataFrame()

    # Normalizar departamento
    df["departamento_limpio"] = df["departamento"].apply(_normalizar_depto)

    # Columnas numéricas
    for nc in ("base_iva", "comision", "costos", "margen"):
        df[nc] = pd.to_numeric(df.get(nc, 0), errors="coerce").fillna(0.0)

    # Total Venta Real
    df["total_venta_real"] = df["base_iva"] + df["comision"]

    # Fechas
    MMAP = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
            7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
    try:
        fechas_dt = pd.to_datetime(df["fecha"], errors="coerce")
        df["fecha"] = fechas_dt.dt.strftime("%Y-%m-%d")
    except Exception:
        fechas_dt = pd.Series([pd.NaT] * len(df), index=df.index)
        df["fecha"] = df["fecha"].astype(str)

    if mes is None:
        df["mes"] = fechas_dt.dt.month.map(MMAP).fillna("SIN MES")
        df["ano"] = fechas_dt.dt.year.fillna(ano or datetime.now().year).astype(int)
    else:
        df["mes"] = mes
        df["ano"] = ano or datetime.now().year

    df["empresa"] = empresa
    df["_fecha_dt"] = fechas_dt

    # Limpiar strings
    for sc in ("cliente","departamento","departamento_limpio","tipo","ciudad"):
        if sc in df.columns:
            df[sc] = df[sc].astype(str).str.strip().replace("nan","").replace("None","")

    return df


def parsear_excel(contenido: bytes) -> pd.DataFrame:
    """
    Parsea el Excel en memoria (bytes).
    Nunca escribe en disco. Retorna DataFrame consolidado.
    """
    xl     = pd.ExcelFile(io.BytesIO(contenido))
    disp   = xl.sheet_names
    ano_hoy = datetime.now().year
    frames = []

    # Hojas mensuales
    for hoja in MESES:
        if hoja in disp:
            df = _leer_hoja(xl, hoja, "Paradais", mes=hoja, ano=ano_hoy)
            if not df.empty:
                frames.append(df)

    # Hojas Paradais Media
    encontradas = set()
    for hoja in EXTRA_SHEETS:
        if hoja in disp:
            df = _leer_hoja(xl, hoja, "Paradais Media", mes=None, ano=ano_hoy)
            if not df.empty:
                frames.append(df)
            encontradas.add(hoja)

    if not encontradas:
        for hoja in disp:
            u = hoja.upper().strip()
            if ("PARADAIS MEDIA" in u or u.startswith("MEDIA")) and hoja not in encontradas:
                df = _leer_hoja(xl, hoja, "Paradais Media", mes=None, ano=ano_hoy)
                if not df.empty:
                    frames.append(df)
                encontradas.add(hoja)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True, sort=False)


# ─────────────────────────────────────────────────────────────────────────────
# GENERADOR DE EXCEL DESCARGABLE
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel_descarga(df: pd.DataFrame, original_bytes: bytes) -> bytes:
    """
    Genera Excel en memoria (BytesIO): hojas originales intactas +
    nueva hoja 'DATA_LIMPIA_Y_ESTRUCTURADA' al inicio.
    NUNCA escribe en disco.
    """
    cols_exp  = [c for c in EXPORT_COLS if c in df.columns]
    hdrs_exp  = [EXPORT_HEADERS[EXPORT_COLS.index(c)] for c in cols_exp]
    df_export = df[cols_exp].copy()
    df_export.columns = hdrs_exp

    # Redondear columnas numéricas
    for col in ("Base IVA","Comisión","Total Venta Real","Costos","Margen"):
        if col in df_export.columns:
            df_export[col] = pd.to_numeric(df_export[col], errors="coerce").round(2)

    # Cargar workbook original
    try:
        wb = openpyxl.load_workbook(io.BytesIO(original_bytes), data_only=True)
    except Exception:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Crear hoja limpia en posición 0
    ws = wb.create_sheet("DATA_LIMPIA_Y_ESTRUCTURADA", 0)

    H_FILL  = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL= PatternFill("solid", fgColor="EFF6FF")
    H_FONT  = Font(bold=True, color="FFFFFF", size=10)
    D_FONT  = Font(size=9)
    CENTER  = Alignment(horizontal="center", vertical="center")

    # Encabezados
    for ci, hdr in enumerate(df_export.columns, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill, c.font, c.alignment = H_FILL, H_FONT, CENTER

    # Datos
    for ri, row in enumerate(df_export.itertuples(index=False), 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = D_FONT
            if fill:
                c.fill = fill

    # Tabla nativa de Excel
    n_filas = len(df_export)
    if n_filas > 0:
        last_col = openpyxl.utils.get_column_letter(len(df_export.columns))
        ref      = f"A1:{last_col}{n_filas + 1}"
        tbl      = Table(displayName="DataLimpia", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tbl)

    # Auto-ancho de columnas
    for col_cells in ws.columns:
        ancho = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(ancho + 3, 36)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE FORMATO
# ─────────────────────────────────────────────────────────────────────────────
def _m(v: float) -> str:
    return f"${v:,.0f}"

def _p(v: float) -> str:
    return f"{v:.1f}%"

_PLOT_LAYOUT = dict(
    paper_bgcolor="#FFFFFF",
    plot_bgcolor="#FFFFFF",
    font=dict(family="Inter, Arial, sans-serif", color="#1C1917", size=12),
    margin=dict(l=0, r=0, t=10, b=0),
    xaxis=dict(showgrid=True, gridcolor="#F5F5F4", zeroline=False,
               tickfont=dict(size=11, color="#78716C")),
    yaxis=dict(showgrid=True, gridcolor="#F5F5F4", zeroline=False,
               tickfont=dict(size=11, color="#78716C")),
)

def _kpi(col, label: str, valor: str, badge_txt: str = "", badge_cls: str = "badge-neutral"):
    col.markdown(
        f'<div class="kpi-card">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{valor}</div>'
        f'<span class="kpi-badge {badge_cls}">{badge_txt}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

def _semaforo(label: str, real: float, meta: float, fmt=_m):
    pct = (real / meta * 100) if meta else 0
    if pct >= 100:
        css, ico = "sem-verde", "✅"
    elif pct >= 85:
        css, ico = "sem-amber", "⚠️"
    else:
        css, ico = "sem-rojo",  "🔴"
    st.markdown(
        f'<div class="{css}"><b>{ico} {label}</b>&nbsp;&nbsp;'
        f'Real: <b>{fmt(real)}</b> &nbsp;·&nbsp; Meta: {fmt(meta)} '
        f'&nbsp;·&nbsp; Cumplimiento: <b>{pct:.1f}%</b></div>',
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RENDERIZADO DEL DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def render_dashboard(df: pd.DataFrame):

    # ── Preparar opciones de filtros ──────────────────────────────────────
    df["_fecha_dt"] = pd.to_datetime(df["fecha"], errors="coerce")
    f_min = df["_fecha_dt"].min()
    f_max = df["_fecha_dt"].max()

    tipos    = ["Todos"] + sorted(df["tipo"].replace("","—").dropna().unique().tolist())
    ciudades = ["Todas"] + sorted(df["ciudad"].replace("","—").dropna().unique().tolist())
    deptos   = ["Todos"] + sorted([d for d in df["departamento_limpio"].dropna().unique() if d])
    clientes = ["Todos"] + sorted([c for c in df["cliente"].dropna().unique() if c])

    # ── HEADER ────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="ddb-header">'
        '<div class="ddb-logo">'
        '<svg viewBox="0 0 44 44" fill="none" xmlns="http://www.w3.org/2000/svg">'
        '<!-- B superior -->'
        '<rect x="10" y="7" width="6" height="13" rx="1" fill="#EAB308"/>'
        '<path d="M16 7 H22 Q28 7 28 13.5 Q28 20 22 20 H16 Z" fill="#EAB308"/>'
        '<!-- B inferior -->'
        '<rect x="10" y="22" width="6" height="15" rx="1" fill="#EAB308"/>'
        '<path d="M16 22 H23 Q30 22 30 29.5 Q30 37 23 37 H16 Z" fill="#EAB308"/>'
        '</svg>'
        '</div>'
        '<div>'
        '<div class="ddb-title">Ventas &amp; Costos</div>'
        '<div class="ddb-sub">Paradais DDB · Dashboard Ejecutivo</div>'
        '</div>'
        '<div class="ddb-badge">🟢 Conectado al Excel en Vivo</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    # ── Barra de filtros horizontal ───────────────────────────────────────
    with st.container():
        st.markdown(
            "<div style='background:#FFFFFF;border:1px solid #E7E5E4;border-radius:12px;"
            "padding:0.75rem 1.2rem 0.5rem 1.2rem;margin-bottom:1rem;"
            "box-shadow:rgba(0,0,0,0.03) 0px 2px 8px;'>"
            "<span style='font-size:0.72rem;font-weight:700;color:#78716C;"
            "text-transform:uppercase;letter-spacing:0.08em'>🔎 Filtros</span>"
            "</div>",
            unsafe_allow_html=True,
        )
        fc1, fc2, fc3, fc4, fc5, fc6, fc7 = st.columns([1.2, 1.2, 1.5, 1.8, 1.1, 1.1, 0.8])
        with fc1:
            sel_tipo    = st.selectbox("Tipo",          tipos,    label_visibility="visible")
        with fc2:
            sel_ciudad  = st.selectbox("Ciudad",        ciudades, label_visibility="visible")
        with fc3:
            sel_depto   = st.selectbox("Departamento",  deptos,   label_visibility="visible")
        with fc4:
            sel_cliente = st.selectbox("Cliente",       clientes, label_visibility="visible")
        if pd.notna(f_min) and pd.notna(f_max):
            with fc5:
                desde = st.date_input("Desde", value=f_min.date(),
                                      min_value=f_min.date(), max_value=f_max.date())
            with fc6:
                hasta = st.date_input("Hasta", value=f_max.date(),
                                      min_value=f_min.date(), max_value=f_max.date())
        else:
            desde = hasta = None
        with fc7:
            st.markdown("<div style='margin-top:1.6rem'>", unsafe_allow_html=True)
            if st.button("🚪 Salir", use_container_width=True):
                for k in ("autenticado", "df", "excel_bytes", "nombre_archivo"):
                    st.session_state.pop(k, None)
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

    # ── Aplicar filtros ────────────────────────────────────────────────────
    fdf = df.copy()
    if sel_tipo    != "Todos":   fdf = fdf[fdf["tipo"]                == sel_tipo]
    if sel_ciudad  != "Todas":   fdf = fdf[fdf["ciudad"]              == sel_ciudad]
    if sel_depto   != "Todos":   fdf = fdf[fdf["departamento_limpio"] == sel_depto]
    if sel_cliente != "Todos":   fdf = fdf[fdf["cliente"]             == sel_cliente]
    if desde and hasta:
        fdf = fdf[
            (fdf["_fecha_dt"].dt.date >= desde) &
            (fdf["_fecha_dt"].dt.date <= hasta)
        ]

    if len(fdf) == 0:
        st.warning("No hay datos para los filtros seleccionados.")
        return

    # ── KPIs ──────────────────────────────────────────────────────────────
    ventas = fdf["total_venta_real"].sum()
    costos = fdf["costos"].sum()
    margen = ventas - costos
    rentab = (margen / ventas * 100) if ventas else 0.0

    k1, k2, k3, k4 = st.columns(4)

    _kpi(k1, "Ventas Totales", _m(ventas),
         "Ingresos consolidados", "badge-neutral")
    _kpi(k2, "Costos Totales", _m(costos),
         "Estructura de costos", "badge-neutral")

    if margen >= 0:
        _kpi(k3, "Utilidad Bruta", _m(margen), "Saludable ✓", "badge-green")
    else:
        _kpi(k3, "Utilidad Bruta", _m(margen), "En riesgo", "badge-red")

    if rentab >= 30:
        _kpi(k4, "% Rentabilidad", _p(rentab), "Óptimo ✓", "badge-green")
    elif rentab >= 15:
        _kpi(k4, "% Rentabilidad", _p(rentab), "En alerta", "badge-amber")
    else:
        _kpi(k4, "% Rentabilidad", _p(rentab), "Crítico", "badge-red")

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # FILA DE GRÁFICOS CENTRALES
    # ══════════════════════════════════════════════════════════════════════
    col_izq, col_der = st.columns([1.1, 0.9])

    # ── A. Tendencia Mensual: Ventas / Costos / Margen (Spline) ──────────
    with col_izq:
        st.markdown('<div class="section-title">📈 Tendencia Mensual · Ventas vs Costos vs Margen</div>',
                    unsafe_allow_html=True)
        df_mes = (
            fdf.groupby("mes").agg(
                Ventas=("total_venta_real", "sum"),
                Costos=("costos", "sum"),
            ).reset_index()
        )
        df_mes["Margen"] = df_mes["Ventas"] - df_mes["Costos"]
        df_mes["_ord"]   = df_mes["mes"].map(MES_NUM).fillna(99)
        df_mes = df_mes.sort_values("_ord")

        if not df_mes.empty:
            fig_a = go.Figure()
            # Ventas — línea dorada con área suave
            fig_a.add_trace(go.Scatter(
                x=df_mes["mes"], y=df_mes["Ventas"],
                name="Ventas", mode="lines+markers",
                line=dict(color="#EAB308", width=2.5, shape="spline"),
                marker=dict(size=6),
                fill="tozeroy",
                fillcolor="rgba(234,179,8,0.05)",
            ))
            # Costos — gris oscuro
            fig_a.add_trace(go.Scatter(
                x=df_mes["mes"], y=df_mes["Costos"],
                name="Costos", mode="lines+markers",
                line=dict(color="#475569", width=2, shape="spline"),
                marker=dict(size=5),
            ))
            # Margen — verde esmeralda discontinuo
            fig_a.add_trace(go.Scatter(
                x=df_mes["mes"], y=df_mes["Margen"],
                name="Margen Neto", mode="lines+markers",
                line=dict(color="#10B981", width=2, dash="dot", shape="spline"),
                marker=dict(size=5),
            ))
            fig_a.update_layout(
                **_PLOT_LAYOUT,
                height=320,
                legend=dict(orientation="h", yanchor="bottom", y=1.02,
                            xanchor="right", x=1, font=dict(size=11)),
            )
            fig_a.update_yaxes(tickprefix="$", tickformat=",.0f")
            st.plotly_chart(fig_a, use_container_width=True)

    # ── B. Dona — Distribución por Departamento con total al centro ───────
    with col_der:
        st.markdown('<div class="section-title">🍩 Participación por Departamento</div>',
                    unsafe_allow_html=True)
        df_dept = (
            fdf.groupby("departamento_limpio")["total_venta_real"]
            .sum().reset_index()
            .rename(columns={"departamento_limpio": "Departamento",
                             "total_venta_real":    "Ventas"})
            .sort_values("Ventas", ascending=False)
        )
        df_dept = df_dept[df_dept["Ventas"] > 0]

        if not df_dept.empty:
            _DONUT_COLORS = [
                "#EAB308","#1C1917","#10B981","#475569",
                "#F59E0B","#DC2626","#6366F1","#0EA5E9",
            ]
            total_str = _m(ventas)
            fig_b = go.Figure(go.Pie(
                labels=df_dept["Departamento"],
                values=df_dept["Ventas"],
                hole=0.62,
                textinfo="percent",
                textfont=dict(size=11, color="#1C1917"),
                marker=dict(colors=_DONUT_COLORS[:len(df_dept)],
                            line=dict(color="#FFFFFF", width=2)),
                hovertemplate="<b>%{label}</b><br>%{value:$,.0f}<br>%{percent}<extra></extra>",
            ))
            fig_b.update_layout(
                **{k: v for k, v in _PLOT_LAYOUT.items() if k not in ("xaxis","yaxis")},
                height=320,
                annotations=[dict(
                    text=f"<b>{total_str}</b>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=17, color="#1C1917",
                              family="Georgia, 'Times New Roman', serif"),
                    xref="paper", yref="paper",
                )],
                legend=dict(orientation="v", font=dict(size=10),
                            x=1.02, y=0.5, yanchor="middle"),
                showlegend=True,
            )
            st.plotly_chart(fig_b, use_container_width=True)

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN INFERIOR — BP + Barras horizontales top 10
    # ══════════════════════════════════════════════════════════════════════
    mask_bp = fdf["cliente"].str.upper().str.strip().str.contains(
        r"BANCO.*PAC[IÍ]FICO|PAC[IÍ]FICO.*BANCO", regex=True, na=False
    )
    df_bp   = fdf[mask_bp]
    df_rest = fdf[~mask_bp]

    st.markdown("<br>", unsafe_allow_html=True)
    bot_izq, bot_der = st.columns([0.9, 1.1])

    # ── Bloque Banco del Pacífico ──────────────────────────────────────────
    with bot_izq:
        v_bp = df_bp["total_venta_real"].sum()
        c_bp = df_bp["costos"].sum()
        m_bp = v_bp - c_bp
        r_bp = (m_bp / v_bp * 100) if v_bp else 0.0

        st.markdown(
            '<div class="bp-block">'
            '<div class="bp-block-title">🏦 Banco del Pacífico</div>',
            unsafe_allow_html=True,
        )
        mb1, mb2 = st.columns(2)
        mb1.metric("Ventas",       _m(v_bp))
        mb2.metric("Costos",       _m(c_bp))
        mb3, mb4 = st.columns(2)
        mb3.metric("Utilidad",     _m(m_bp))
        mb4.metric("Rentabilidad", _p(r_bp))

        if v_bp > 0:
            st.markdown(
                "<div style='font-size:0.78rem;font-weight:700;"
                "color:#78716C;margin:0.8rem 0 0.4rem 0;"
                "text-transform:uppercase;letter-spacing:0.06em'>"
                "Semáforos de cumplimiento</div>",
                unsafe_allow_html=True,
            )
            _semaforo("Ventas BP",  v_bp, v_bp * 0.90)
            _semaforo("Margen BP",  m_bp, m_bp * 0.85 if m_bp > 0 else 1.0)
        else:
            st.info("Sin datos de Banco del Pacífico.")

        st.markdown("</div>", unsafe_allow_html=True)

    # ── C. Barras horizontales Top 10 (sin BP) ─────────────────────────────
    with bot_der:
        st.markdown(
            '<div class="section-title">📊 Top 10 Clientes · Resto de Agencia '
            '<span style="font-size:0.72rem;font-weight:400;color:#78716C">'
            '(excl. Banco del Pacífico)</span></div>',
            unsafe_allow_html=True,
        )
        v_r = df_rest["total_venta_real"].sum()
        top10 = (
            df_rest.groupby("cliente")["total_venta_real"]
            .sum().sort_values(ascending=False).head(10).reset_index()
        )
        top10.columns = ["Cliente", "Ventas"]
        top10 = top10.sort_values("Ventas", ascending=True)  # plotly invierte

        if not top10.empty:
            # Semáforo por participación relativa
            def _bar_color(v):
                pct = (v / v_r * 100) if v_r else 0
                if pct >= 15: return "#10B981"
                if pct >= 8:  return "#F59E0B"
                return "#DC2626"

            colors = [_bar_color(v) for v in top10["Ventas"]]

            fig_c = go.Figure(go.Bar(
                x=top10["Ventas"],
                y=top10["Cliente"],
                orientation="h",
                marker=dict(color=colors, line=dict(color="rgba(0,0,0,0)")),
                text=[_m(v) for v in top10["Ventas"]],
                textposition="outside",
                textfont=dict(size=10, color="#1C1917"),
                hovertemplate="<b>%{y}</b><br>%{x:$,.0f}<extra></extra>",
            ))
            fig_c.update_layout(
                **{k: v for k, v in _PLOT_LAYOUT.items() if k not in ("xaxis", "yaxis")},
                height=max(280, len(top10) * 34),
                yaxis=dict(tickfont=dict(size=10, color="#1C1917"),
                           showgrid=False, zeroline=False),
                xaxis=dict(tickprefix="$", tickformat=",.0f",
                           showgrid=True, gridcolor="#F5F5F4"),
            )
            st.plotly_chart(fig_c, use_container_width=True)

    # ── Tabla detalle (expandible) ─────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("📋 Ver datos filtrados", expanded=False):
        cols_vis = [c for c in EXPORT_COLS if c in fdf.columns]
        df_vis   = fdf[cols_vis].copy()
        df_vis.columns = [EXPORT_HEADERS[EXPORT_COLS.index(c)] for c in cols_vis]
        st.dataframe(df_vis, use_container_width=True, height=380)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
_MAX_MB   = 50
_EXTS_OK  = {".xlsx", ".xlsm"}
_MAGIC_PK = b"PK\x03\x04"    # firma ZIP/OOXML


def main():
    # ── Autenticación ──────────────────────────────────────────────────────
    if not pantalla_login():
        st.stop()

    st.markdown(
        "<h2 style='color:#e2e8f0;margin-bottom:0'>📊 Paradais DDB · Dashboard Ejecutivo</h2>"
        "<p style='color:#94a3b8;margin-top:0.2rem'>Carga tu archivo Excel para visualizar los datos financieros.</p>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    # ── Si ya hay datos en sesión → mostrar dashboard ──────────────────────
    if st.session_state.get("df") is not None:
        df = st.session_state["df"]

        col_dl, col_nuevo, _ = st.columns([2, 2, 4])
        with col_dl:
            st.download_button(
                label="⬇️  Descargar Excel Limpio",
                data=st.session_state["excel_bytes"],
                file_name=f"DATA_LIMPIA_{st.session_state.get('nombre_archivo','export')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_nuevo:
            if st.button("📂 Cargar nuevo archivo", use_container_width=True):
                for k in ("df", "excel_bytes", "nombre_archivo"):
                    st.session_state.pop(k, None)
                st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        render_dashboard(df)
        return

    # ── Carga de archivo ───────────────────────────────────────────────────
    subido = st.file_uploader(
        "Selecciona el archivo Excel de ventas",
        type=["xlsx", "xlsm"],
        help=f"Máximo {_MAX_MB} MB · Formatos: .xlsx · .xlsm",
    )

    if subido is None:
        st.info("⬆️  Carga un archivo .xlsx o .xlsm para comenzar.")
        return

    # Validación de extensión
    ext = ("." + subido.name.rsplit(".", 1)[-1].lower()) if "." in subido.name else ""
    if ext not in _EXTS_OK:
        st.error("Error en el formato del archivo cargado.")
        return

    # Leer bytes
    contenido = subido.read()

    # Validación de tamaño
    if len(contenido) / (1024 * 1024) > _MAX_MB:
        st.error(f"El archivo supera el límite de {_MAX_MB} MB.")
        return

    # Validación de magic bytes (firma OOXML/ZIP)
    if contenido[:4] != _MAGIC_PK:
        st.error("Error en el formato del archivo cargado.")
        return

    # ── Procesamiento en memoria ───────────────────────────────────────────
    with st.spinner("⚙️  Procesando datos, por favor espera…"):
        try:
            df = parsear_excel(contenido)

            if df.empty:
                st.warning(
                    "No se encontraron datos válidos. Verifica que el archivo contenga "
                    "hojas mensuales (ENERO–DICIEMBRE) o 'Paradais Media 2026'."
                )
                return

            excel_bytes = generar_excel_descarga(df, contenido)

            # Guardar SOLO en st.session_state (RAM volátil de la sesión)
            # Nunca se escribe en disco.
            st.session_state["df"]             = df
            st.session_state["excel_bytes"]    = excel_bytes
            st.session_state["nombre_archivo"] = subido.name.rsplit(".", 1)[0]

        except Exception:
            # No exponer trazas de error al usuario
            st.error("Error en el formato del archivo cargado.")
            return

    st.success(f"✅  {len(df):,} registros procesados correctamente.")
    st.rerun()


if __name__ == "__main__":
    main()
